import json
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, output_dir: str = "results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_results(
        self,
        evaluated_students: list,
        rejected_students: list,
        filename: str = "baremo_resultados.xlsx",
    ) -> str:
        wb = Workbook()

        self._write_aptos(wb, evaluated_students)
        self._write_no_aptos(wb, evaluated_students)
        self._write_resumen(wb, evaluated_students, rejected_students)

        output_path = self.output_dir / filename
        wb.save(str(output_path))
        logger.info(f"Excel generado: {output_path}")
        return str(output_path)

    def _write_aptos(self, wb: Workbook, students: list):
        ws = wb.active
        ws.title = "Aptos"

        aptos = [s for s in students if s.get("apto")]
        req_keys = self._get_requisito_keys(students)

        headers = ["Orden", "ID Alumno", "Estado", "Descripción"]
        for rk in req_keys:
            headers.append(f"Requisito: {rk}")

        self._style_header(ws, headers)

        for i, s in enumerate(aptos, 1):
            row = i + 1
            col = 1
            ws.cell(row=row, column=col, value=s.get("orden", i)); col += 1
            ws.cell(row=row, column=col, value=s.get("student_id", "")); col += 1
            ws.cell(row=row, column=col, value=s.get("estado", "")); col += 1
            ws.cell(row=row, column=col, value=s.get("descripcion", "")); col += 1
            reqs = {r["clave"]: r for r in s.get("requisitos", [])}
            for rk in req_keys:
                r = reqs.get(rk, {})
                val = "✓" if r.get("cumple") else "✗"
                ws.cell(row=row, column=col, value=val); col += 1

        self._auto_width(ws, headers)

    def _write_no_aptos(self, wb: Workbook, students: list):
        ws = wb.create_sheet("No aptos")

        no_aptos = [s for s in students if not s.get("apto")]
        req_keys = self._get_requisito_keys(students)

        headers = ["Orden", "ID Alumno", "Estado", "Descripción"]
        for rk in req_keys:
            headers.append(f"Requisito: {rk}")

        self._style_header(ws, headers)

        for i, s in enumerate(no_aptos, 1):
            row = i + 1
            col = 1
            ws.cell(row=row, column=col, value=s.get("orden", i)); col += 1
            ws.cell(row=row, column=col, value=s.get("student_id", "")); col += 1
            ws.cell(row=row, column=col, value=s.get("estado", "")); col += 1
            ws.cell(row=row, column=col, value=s.get("descripcion", "")); col += 1
            reqs = {r["clave"]: r for r in s.get("requisitos", [])}
            for rk in req_keys:
                r = reqs.get(rk, {})
                val = "✓" if r.get("cumple") else "✗"
                ws.cell(row=row, column=col, value=val); col += 1

        self._auto_width(ws, headers)

    def _get_requisito_keys(self, students: list) -> list[str]:
        keys: set[str] = set()
        for s in students:
            for r in s.get("requisitos", []):
                keys.add(r.get("clave", ""))
        return sorted(k for k in keys if k)

    def _write_resumen(self, wb: Workbook, evaluados: list, descartados: list):
        ws = wb.create_sheet("Resumen")

        ws.cell(row=1, column=1, value="Métrica").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Valor").font = Font(bold=True)

        ws.cell(row=2, column=1, value="Total Alumnos")
        ws.cell(row=2, column=2, value=len(evaluados))

        aptos = sum(1 for s in evaluados if s.get("apto"))
        no_aptos = len(evaluados) - aptos

        ws.cell(row=3, column=1, value="Aptos")
        ws.cell(row=3, column=2, value=aptos)

        ws.cell(row=4, column=1, value="No aptos")
        ws.cell(row=4, column=2, value=no_aptos)

        if evaluados and evaluados[0].get("requisitos"):
            ws.cell(row=6, column=1, value="Desglose por requisito").font = Font(bold=True)
            row_num = 7
            req_keys = self._get_requisito_keys(evaluados)
            for rk in req_keys:
                total = len(evaluados)
                cumplen = sum(1 for s in evaluados if any(
                    r.get("clave") == rk and r.get("cumple") for r in s.get("requisitos", [])
                ))
                pct = (cumplen / total * 100) if total > 0 else 0
                ws.cell(row=row_num, column=1, value=rk)
                ws.cell(row=row_num, column=2, value=f"{cumplen}/{total} ({pct:.0f}%)")
                row_num += 1

        self._auto_width(ws, ["Métrica", "Valor"])

    def _style_header(self, ws, headers: list[str]):
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws, headers: list[str]):
        for col, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 3, 18)
